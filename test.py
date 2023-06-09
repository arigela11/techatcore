#openpyxl is a python library created  to read or write excel files

import openpyxl

def update_records_from_excel(excel_file_path):
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        
        # Assuming the data is in the first sheet
        sheet = workbook.active
        #to print the table:
        # Iterate over the rows, starting from the second row (excluding header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0]  # Assuming the first column contains the names
            weight_value = row[1]  # Assuming the second column contains their weights
            print(f"before updating{name},weight{weight_value}")
            # Perform the record update logic here for bulk records

            # Connect to the database and update the record with the given ID using the update value
           

            print(f"Updated record with name {name} to {weight_value}")
        #To update individually  
        sheet['B5']=60
        print(sheet['B5'].value) #another method(print(sheet.cell(row=5,column=2).value)
       
        #To add new values:
        sheet['A7']="oopslu"
        sheet["B7"]=11
        sheet["A8"]="ravi"
        sheet["B8"]=78
        #to reflect in the excel file
        workbook.save(excel_file_path)

        
        print("All records updated successfully.")
    
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Prompt the manager for the Excel file path
excel_file_path = input("Please provide the path to the Excel file: ")
#excel_file_path="C:\\Users\\student\\Desktop\\test.xlsx"
# Call the function to update the records
update_records_from_excel(excel_file_path)
